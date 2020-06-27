# -*- coding: utf8 -*-

import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill
import math
import pywt
import csv

# CONFIGURAÇÕES - ARQUIVO EXCEL
# A -> W+RFC            B -> W+1NN          C -> W+GNB          D -> W + SVM
# E -> W+PCA+RFC        F -> W+PCA+1NN      G -> W+PCA+GNB      H -> W+PCA+SVM
# I -> W+LDA+RFC        J -> W+LDA+1NN      K -> W+LDA+GNB      L -> W+LDA+SVM
# M -> W+PCA+LDA+RFC    N -> W+PCA+LDA+1NN  O -> W+PCA+LDA+GNB  P -> W+PCA+LDA+SVM

if __name__ == "__main__":
    workbooks = ["C:\\Users\\fabri\\Python Projects\\Waveletfaces Parametrizaveis\\RESULTADOS\\1 Dimension\\Overall Accuracy per Wavelet\\CSVs\\AR_1D_Parametrization_lvl_3_pca_False_lda_True.csv",
                 "C:\\Users\\fabri\\Python Projects\\Waveletfaces Parametrizaveis\\RESULTADOS\\1 Dimension\\Overall Accuracy per Wavelet\\CSVs\\YaleB_1D_Parametrization_lvl_4_pca_False_lda_True.csv",
                 "C:\\Users\\fabri\\Python Projects\\Waveletfaces Parametrizaveis\\RESULTADOS\\1 Dimension\\Overall Accuracy per Wavelet\\CSVs\\ORL_1D_Parametrization_lvl_2_pca_True_lda_True.csv",
                 "C:\\Users\\fabri\\Python Projects\\Waveletfaces Parametrizaveis\\RESULTADOS\\1 Dimension\\Overall Accuracy per Wavelet\\CSVs\\GTech_1D_Parametrization_lvl_5_pca_False_lda_True.csv",
                 "C:\\Users\\fabri\\Python Projects\\Waveletfaces Parametrizaveis\\RESULTADOS\\1 Dimension\\Overall Accuracy per Wavelet\\CSVs\\Essex_1D_Parametrization_lvl_5_pca_True_lda_True.csv"]
    
    blackFill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    grayFill = PatternFill(start_color='A8A8A8', end_color='A8A8A8', fill_type="solid")

    for i in range(0, len(workbooks)):
        count, classifier, config = 0, 0, -1
        mean, std = 0, 0
        
        if i == 0:    best_mean, best_std = 97.76925, 0.394069445                # AR - Level 3 - W+LDA+1NN                   
        elif i == 1:  best_mean, best_std = 91.9087, 0.764538207                 # YaleB - Level 4 - W+LDA+1NN
        elif i == 2:  best_mean, best_std = 97.58, 1.040961094                   # ORL - Level 2 - W+PCA+LDA+1NN        HIGHEST ACCURACY OF WAVELETS
        elif i == 3:  best_mean, best_std = 83.885, 1.645227948                  # GTech - Level 5 - W+LDA+1NN 
        elif i == 4:  best_mean, best_std = 94.94997, 0.848323434                # Essex - Level 5 - W+PCA+LDA+1NN 
        
        wb = load_workbook(filename=workbooks[i])
        wb2 = Workbook()
        
        for level in levels:
            wavelet = ""
            sign = False
            table = wb2.create_sheet(title=databases[i] + "_" + level)
            sheet = wb[databases[i] + "_" + level]

            for row in sheet.rows:
                for cell in row:
                    if cell.value == None or cell.value == "-": continue
                    count += 1
                    if count <= 20: continue
                    if isinstance(cell.value, basestring):
                        if cell.value == wavelet: config += 1
                        else: config = 0
                        wavelet = cell.value
                        classifier = 0
                    else:
                        if sign == True:
                            sign = False
                            std = cell.value
                            interval_max = (best_mean - mean) + 1.96 * math.sqrt(best_std ** 2 + std ** 2)
                            interval_min = (best_mean - mean) - 1.96 * math.sqrt(best_std ** 2 + std ** 2)
                            if interval_min <= 0 <= interval_max:       # Does the interval contain 0?
                                if config == 0:
                                    if classifier == 0: col = "A"
                                    if classifier == 1: col = "B"
                                    if classifier == 2: col = "C"
                                    if classifier == 3: col = "D"
                                elif config == 1:
                                    if classifier == 0: col = "E"
                                    if classifier == 1: col = "F"
                                    if classifier == 2: col = "G"
                                    if classifier == 3: col = "H"
                                elif config == 2:
                                    if classifier == 0: col = "I"
                                    if classifier == 1: col = "J"
                                    if classifier == 2: col = "K"
                                    if classifier == 3: col = "L"
                                elif config == 3:
                                    if classifier == 0: col = "M"
                                    if classifier == 1: col = "N"
                                    if classifier == 2: col = "O"
                                    if classifier == 3: col = "P"
                                table[col + str(waves.index(wavelet) + 1)].fill = grayFill
                            classifier += 1
                        else:
                            mean = cell.value                    
                            sign = True
        wb2.save(databases[i] + ".xlsx")                
